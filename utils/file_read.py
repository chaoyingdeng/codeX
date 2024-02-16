from pathlib import Path
from basic.exceptions import FileNotExists
from basic.exceptions import ExcelSheetError
from yaml import safe_load_all, safe_load
from openpyxl import load_workbook
import configparser


class File:
    def __init__(self, files_path):
        if not Path(files_path).exists():
            raise FileNotExists(files_path)

        self._file_path = files_path
        self._data = None


class YamlReader(File):
    """ multi指定是否进行多节读取 """
    def __init__(self, yml_path, multi: bool = False):
        super().__init__(yml_path)
        self._multi = multi
        self._data = self._load_data()

    def _load_data(self):
        with open(self._file_path, 'rb') as fp:
            return list(safe_load_all(fp)) if self._multi else safe_load(fp)

    @property
    def data(self):
        return self._data

    # 不支持多节
    def __getattr__(self, item):
        return tuple(self._data.get(item, None)) if not self._multi else None


class ExcelReader(File):
    """ 默认不要表头,把每行数据转成一个元组放置到列表中 """
    def __init__(self, excel_path, sheet=0):
        if not isinstance(sheet, int):
            raise ExcelSheetError(sheet)
        super().__init__(excel_path)
        self._sheet = sheet
        self._data = self._load_data()

    def _load_data(self):
        work_book = load_workbook(self._file_path).worksheets[self._sheet]
        return [cell for cell in work_book.iter_rows(min_row=2, values_only=True)]

    @property
    def data(self):
        return self._data


class IniReader(File):
    def __init__(self, ini_path):
        super().__init__(ini_path)
        self._cfg = configparser.ConfigParser()
        self._cfg.read(ini_path, encoding='utf-8')
        self._data = self._load_data()

    def _load_data(self):
        all_options = {}
        for section in self._cfg.sections():
            for option in self._cfg.options(section):
                value = self._cfg.get(section, option)
                all_options[option] = value
        return all_options

    @property
    def data(self):
        return self._data

    def __getattr__(self, item):
        return self._data.get(item, None)


if __name__ == '__main__':
    file_path = r'/Users/dengchaoying/PycharmProjects/CodeX/business/sfe/sfe_locators.yaml'
    yml = YamlReader(file_path)